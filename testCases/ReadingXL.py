import openpyxl

path = "C://Users//Nani Madhan//OneDrive//Desktop//AutomationData//testxl.xlsx"

Workbook = openpyxl.load_workbook(path)

sheet = Workbook.active             #workbook.get_sheet_by_name

rows = sheet.max_row
columns = sheet.max_column

for i in range(1, rows+1):
    for j in range(1, columns+1):
        print(sheet.cell(row=i,column=j).value,end = "  ")
    print()

